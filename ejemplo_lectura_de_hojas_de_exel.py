import openpyxl
import os

workbook = openpyxl.load_workbook('pathfile')
type(workbook)

sheet = workbook.get_sheet_by_name('sheetname')
workbook.get_sheet_names()

cell = sheet[
str(cell.value)

sheet.cell(row=1, column=2)

for i in range(1,8):
    print(i,sheet.cell(row=i, column=2).value)
